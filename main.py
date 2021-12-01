import csv
import os
import sqlite3
from icecream import ic

from openpyxl import load_workbook

"""
Устройство базы данных simdatabase.db

Таблицы:
1. Таблица СИМ карт МТС по которым есть информация (mts_current), что они используются (От Пищулина и т.п.).
    - номер телефона (num_tel)
    - номер симкарты (num_sim) (серийный номер)
    - назначение симкарты (purpose) (ГЛОНАСС/GPS АРМ, ГЛОНАСС/GPS АВТОМОБИЛИ, ГЛОНАСС/GPS ЗАПАСНЫЕ,
        СЕТЕВОЙ МОНИТОРИНГ, ОПРОС CSD СЧЕТЧИКОВ, CSD НА СЧЕТЧИКАХ, LORA, Телемеханика, LORA/Телемеханика запасные,
        Личный кабинет, Интернет)
    - Адрес, № ТП, населённый пункт, Точка учёта и т. д. (set_addr) (б-р 50 лет Октября д.14,
                                                            г. Брянск пер Пилотов    Ввод 2,
                                                            Брянск г, Вокзальная ул, дом № 134)
    - Дополнительная информация (aux) (ОДПУ - МУП БрянскГорВодоканал 2019, Юр. лица - 1 - ИП Алхимова,
                                    Юр лица 2020 - Брянск - Храм в честь Воскресения Христова)
    - Серийный номер устройства (s_num) (№ прибора учёта и тп)
    - Тип устройства (type) (Модем, Устройство ГЛОНАСС, Системы безопасности и мониторинга,
                            Автоматические ворота и шлагбаумы, Прочее)
        
2. Таблица СИМ карт МЕГАФОН (megafon_current) по которым есть информация, что они используются (От Пищулина и т.п.).
    - номер телефона (num_tel)
    - номер симкарты (num_sim) (серийный номер)
    - назначение симкарты (purpose) (ГЛОНАСС/GPS АРМ, ГЛОНАСС/GPS АВТОМОБИЛИ, ГЛОНАСС/GPS ЗАПАСНЫЕ,
        СЕТЕВОЙ МОНИТОРИНГ, ОПРОС CSD СЧЕТЧИКОВ, CSD НА СЧЕТЧИКАХ, LORA, Телемеханика, LORA/Телемеханика запасные,
        Личный кабинет, Интернет)
    - Адрес, № ТП, населённый пункт, Точка учёта и т. д. (set_addr) (б-р 50 лет Октября д.14,
                                                            г. Брянск пер Пилотов    Ввод 2,
                                                            Брянск г, Вокзальная ул, дом № 134)
    - Дополнительная информация (aux) (ОДПУ - МУП БрянскГорВодоканал 2019, Юр. лица - 1 - ИП Алхимова,
                                    Юр лица 2020 - Брянск - Храм в честь Воскресения Христова)
    - Серийный номер устройства (s_num) (№ прибора учёта и тп)
    - Тип устройства (type) (Модем, Устройство ГЛОНАСС, Системы безопасности и мониторинга,
                            Автоматические ворота и шлагбаумы, Прочее)

3. Таблица СИМ карт ТЕЛЕ2  (tele2_current)по которым есть информация, что они используются (От Пищулина и т.п.).
    - номер телефона (num_tel)
    - номер симкарты (num_sim) (серийный номер)
    - назначение симкарты (purpose) (ГЛОНАСС/GPS АРМ, ГЛОНАСС/GPS АВТОМОБИЛИ, ГЛОНАСС/GPS ЗАПАСНЫЕ,
        СЕТЕВОЙ МОНИТОРИНГ, ОПРОС CSD СЧЕТЧИКОВ, CSD НА СЧЕТЧИКАХ, LORA, Телемеханика, LORA/Телемеханика запасные,
        Личный кабинет, Интернет)
    - Адрес, № ТП, населённый пункт, Точка учёта и т. д. (set_addr) (б-р 50 лет Октября д.14,
                                                            г. Брянск пер Пилотов    Ввод 2,
                                                            Брянск г, Вокзальная ул, дом № 134)
    - Дополнительная информация (aux) (ОДПУ - МУП БрянскГорВодоканал 2019, Юр. лица - 1 - ИП Алхимова,
                                    Юр лица 2020 - Брянск - Храм в честь Воскресения Христова)
    - Серийный номер устройства (s_num) (№ прибора учёта и тп)
    - Тип устройства (type) (Модем, Устройство ГЛОНАСС, Системы безопасности и мониторинга,
                            Автоматические ворота и шлагбаумы, Прочее)

4. Таблица СИМ карт МТС (mts_operator)
    - номер телефона (num_tel)
    - номер симкарты (num_sim)
    232303055372 charge_report_electric_energy_meters_15_06_2021.xlsx
    232304269877 charge_report_IoT_gprs_sms_call_23_07_2021.xlsx
    232302436150 charge_report_gps_23_07_2021.xlsx
    232304057221 charge_report_management_23_07_2021.xlsx
    232303098297 charge_report_request_for_electricity_meters_23_07_2021.xlsx
    

5. Таблица СИМ карт МЕГАФОН (megafon_operator)
    - номер телефона (num_tel)
    - номер симкарты (num_sim)

6. Таблица СИМ карт ТЕЛЕ2 (tele2_operator)
    - номер телефона (num_tel)
    - номер симкарты (num_sim)
    
--   
select * from mts_operator where num_tel IN
(select num_tel from mts_operator where date = '2021-08-13'
except
select num_tel from mts_current where date = '2021-06-15'); 2332 шт
--
select * from mts_operator where num_tel IN
(select num_tel from mts_operator where date = '2021-08-13'
except
select num_tel from mts_current where date = '2021-06-15')
and date = '2021-08-13';                                    1316 шт
--    
select num_tel from mts_current where date = '2021-06-15';  7674 шт
--
select distinct(num_tel) from mts_operator where num_tel IN
(select num_tel from mts_operator where date = '2021-08-13'
except
select num_tel from mts_current where date = '2021-06-15'); 1316 шт
--
select num_tel from mts_operator where date = '2021-08-13'; 8971 шт 
--
select distinct(num_tel) from mts_operator where date = '2021-06-15'; 8671 шт
--
select distinct(num_tel) from mts_operator where date = '2021-08-13'; 8971 шт 
--
select num_tel from mts_operator where date = '2021-08-13'
except
select num_tel from mts_operator where date = '2021-06-15'; 300 шт
    
"""

def make_db():
    conn = sqlite3.connect("simdatabase.db")
    cursor = conn.cursor()
    cursor.execute("""CREATE TABLE IF NOT EXISTS mts_current
                      (s_id INTEGER PRIMARY KEY,
                      num_tel INTEGER NOT NULL,
                      num_sim INTEGER,
                      purpose TEXT,
                      set_addr TEXT,
                      aux TEXT,
                      s_num TEXT,
                      type TEXT,
                      date TEXT)
                   """)
    cursor.execute("""CREATE TABLE IF NOT EXISTS megafon_current
                      (s_id INTEGER PRIMARY KEY,
                      num_tel INTEGER NOT NULL,
                      num_sim INTEGER,
                      purpose TEXT,
                      set_addr TEXT,
                      aux TEXT,
                      s_num TEXT,
                      type TEXT,
                      date TEXT)
                   """)
    cursor.execute("""CREATE TABLE IF NOT EXISTS tele2_current
                      (s_id INTEGER PRIMARY KEY,
                      num_tel INTEGER NOT NULL,
                      num_sim INTEGER,
                      purpose TEXT,
                      set_addr TEXT,
                      aux TEXT,
                      s_num TEXT,
                      type TEXT,
                      date TEXT)
                   """)
    cursor.execute("""CREATE TABLE IF NOT EXISTS mts_operator
                      (s_id INTEGER PRIMARY KEY,
                      num_tel INTEGER NOT NULL,
                      num_sim INTEGER,
                      account INTEGER,
                      date TEXT)
                   """)
    cursor.execute("""CREATE TABLE IF NOT EXISTS megafon_operator
                      (s_id INTEGER PRIMARY KEY,
                      num_tel INTEGER NOT NULL,
                      num_sim INTEGER,
                      account INTEGER,
                      date TEXT)
                   """)
    cursor.execute("""CREATE TABLE IF NOT EXISTS tele2_operator
                      (s_id INTEGER PRIMARY KEY,
                      num_tel INTEGER NOT NULL,
                      num_sim INTEGER,
                      account INTEGER,
                      date TEXT)
                   """)
    cursor.close()
    conn.close()

READFIRSTMTS: bool = False
READFIRSTMEGAFON: bool = False
READFIRSTMTS210813: bool = False
READFIRSTMTS210910: bool = False
READFIRSTMTS211013: bool = False
READFIRSTMTS211027: bool = False

def mts_site_to_db(excel_filename:str, sheet: str, db_filename: str, columns_to_readwrite: list,
                   date: str = '2021-06-15', log: bool = True):
    wb = load_workbook(excel_filename)
    #print(wb.sheetnames)
    sheet = wb['Charges']
    #print(sheet.title)
    col = 1
    opsos = 'MTS'
    for i in range(2, 11500):
        num = sheet.cell(row=i, column=1).value
        if num:
            row = [col, num, date, opsos, ]
            for c in columns_to_readwrite:
                tval = sheet.cell(row=i, column=c).value
                if tval:
                    val = tval
                row.append(tval)
            col += 1
            conn = sqlite3.connect(db_filename)
            cursor = conn.cursor()
            sql = f"""INSERT INTO mts_operator(num_tel, account, date) VALUES ({row[1]}, {row[4]}, "{row[2]}")"""
            if log:
                ic(sql)
                ic(row)
            else:
                cursor.execute(sql)
                conn.commit()
    cursor.close()
    conn.close()

def mts_site_charge_report_electric_energy_meters_15_06_2021_to_csv():
    excel_filename = "charge_report_electric_energy_meters_15_06_2021.xlsx"
    wb = load_workbook(excel_filename)
    sheet = wb['Charges']
    col = 1
    page = 1
    for i in range(2, 11500):
        file = f'sim-piramida-asque-meter_15_06_2021_{page}.csv'
        if col == 1:
            csv_file = open(file, "w", newline='', encoding='utf-8')
            writer = csv.writer(csv_file, delimiter=';')
            writer.writerow(['Наименование устройства',
                             'Идентификационный номер',
                             'Тип устройства',
                             'Номер телефона',
                             'Адрес, где находится устройство'])
        t_num = sheet.cell(row=i, column=1).value
        if t_num:
            num = t_num[1:]
            auxiliary = f"Счетчик_{page}_{col}"
            snum = ""
            addr = "Г Брянск Станке Димитрова 5В АСКУЭ"
            csv_row = [auxiliary, snum, 'Модем', num, addr]
            print(col, csv_row)
            writer.writerow(csv_row)
            col += 1
        if col == 900:
            col = 1
            page += 1
            csv_file.close()
    csv_file.close()


def mts_site_now_minus_previous_to_csv(now: str = "2021-09-10", previous_date: str = "2021-08-13",
                                       db_filename: str = "simdatabase.db", start_count: int =10000):
    conn = sqlite3.connect(db_filename)
    type = "Модем"
    t_auxiliary: str = "Счетчик"
    addr = "г. Брянск, пр-т Станке Димтрова, 5в, АСКУЭ"
    snum = ""
    num = ""
    cwd = os.getcwd()
    #print(cwd)
    os.chdir("C:/Users/shamin.a/PycharmProjects/simcard")
    file = f"sim-mts_site_now_{now.replace('-', '')}_minus_previous_{previous_date.replace('-', '')}.csv"
    cursor = conn.cursor()
    sql = f"""select distinct(num_tel) from mts_operator where num_tel IN
(select num_tel from mts_operator where date = "{now}"
except
select num_tel from mts_operator where date = "{previous_date}");"""
    print(sql)
    cursor.execute(sql)
    data = cursor.fetchall()
    print(data)
    conn.commit()
    cursor.close()
    conn.close()
    with open(file, "w", newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file, delimiter=';')
        writer.writerow(['Наименование устройства',
                         'Идентификационный номер',
                         'Тип устройства',
                         'Номер телефона',
                         'Адрес, где находится устройство'])
        for i, t_num in enumerate(data):
            num = str(t_num[0])[1:]
            snum = start_count + i
            auxiliary = f"{t_auxiliary}_{snum}"
            row = [auxiliary, snum, type, num, addr]
            ic(row)
            writer.writerow(row)

def mts_site210813_minus_current210615_to_csv():
    conn = sqlite3.connect("simdatabase.db")
    type = "Прочее"
    t_auxiliary = "МТС SIM резерв"
    addr = "г. Брянск, пр-т Станке Димтрова, 5в, АСКУЭ"
    snum = ""
    num = ""
    cwd = os.getcwd()
    # print(cwd)
    os.chdir("C:/Users/shamin.a/PycharmProjects/simcard")
    file = 'sim-mts_site210813_minus_current210615.csv'
    cursor = conn.cursor()
    sql = f"""select distinct(num_tel) from mts_operator where num_tel IN
(select num_tel from mts_operator where date = '2021-08-13'
except
select num_tel from mts_current where date = '2021-06-15');  """
    print(sql)
    cursor.execute(sql)
    data = cursor.fetchall()
    conn.commit()
    cursor.close()
    conn.close()
    with open(file, "w", newline='', encoding='utf-8') as csv_file:
         writer = csv.writer(csv_file, delimiter=';')
         writer.writerow(['Наименование устройства',
                          'Идентификационный номер',
                          'Тип устройства',
                          'Номер телефона',
                          'Адрес, где находится устройство'])
         for i, t_num in enumerate(data):
               num = str(t_num[0])[1:]
               auxiliary = f"{t_auxiliary} {i}"
               row = [auxiliary, snum, type, num, addr]
               print(row)
               writer.writerow(row)

def mts_on_piramida_server_to_csv_db(excel_filename: str = "", db_filename: str = "", db_table:str = "",
                                     columns_to_readwrite: list = ["", ]):
    date = "2021-06-15"
    purpose = "CSD НА СЧЕТЧИКАХ"
    type = "Модем"
    cwd = os.getcwd()
    #print(cwd)
    os.chdir("C:/Users/shamin.a/PycharmProjects/simcard")
    file = 'sim-piramida-asque-meter_15_06_2021.csv'
    wb = load_workbook('SIM-карты ГЛОНАСС АСКУЭ LoRa.xlsx')
    #print(wb.sheetnames)
    sheet = wb['АСКУЭ. Пирамида']
    #print(sheet.title)

    col = 0
    with open(file, "w", newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file, delimiter=';')
        writer.writerow(['Наименование устройства',
                         'Идентификационный номер',
                         'Тип устройства',
                         'Номер телефона',
                         'Адрес, где находится устройство'])
        for i in range(2, 8500):
            num = sheet.cell(row=i, column=5).value
            if num:
                taddr = sheet.cell(row=i, column=2).value
                if taddr: addr = taddr
                t_auxiliary = sheet.cell(row=i, column=3).value
                if t_auxiliary: auxiliary = t_auxiliary
                snum = sheet.cell(row=i, column=4).value
                num = str(num)[1:]
                col += 1
                row = [auxiliary, snum, 'Модем', num, addr]
                writer.writerow(row)
                conn = sqlite3.connect(db_filename)
                cursor = conn.cursor()
                sql = f"""INSERT INTO {db_table}(num_tel, purpose, set_addr, aux, type, s_num, date)
                VALUES (7{row[3]}, '{purpose}', '{addr}', '{auxiliary}', '{row[2]}', '{row[1]}', '{date}')"""
                print(sql)
                cursor.execute(sql)
                print(f"{col} {row}")
                conn.commit()
    cursor.close()
    conn.close()

def mts_on_piramida_server_to_db(now: str = "2021-11-25", excel_filename: str = "SIM-карты АСКУЭ Пирамида 211125.xlsm",
                                 db_filename: str = "simdatabase.db",
                                 db_table:str = "mts_current",
                                 log: bool = True,
                                 columns_to_readwrite: list = ["", ]):
    date = now
    purpose = "CSD НА СЧЕТЧИКАХ"
    type = "Модем"
    cwd = os.getcwd()
    #print(cwd)
    os.chdir("C:/Users/shamin.a/PycharmProjects/simcard")
    file = 'sim-piramida-asque-meter_15_06_2021.csv'
    wb = load_workbook(excel_filename)
    #print(wb.sheetnames)
    sheet = wb['АСКУЭ. Пирамида']
    #print(sheet.title)

    col = 0
    with open(file, "w", newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file, delimiter=';')
        #writer.writerow(['Наименование устройства',
        #                 'Идентификационный номер',
        #                 'Тип устройства',
        #                 'Номер телефона',
        #                 'Адрес, где находится устройство'])
        for i in range(2, 10000):
            num = sheet.cell(row=i, column=5).value
            if num:
                taddr = sheet.cell(row=i, column=2).value
                if taddr: addr = taddr
                t_auxiliary = sheet.cell(row=i, column=3).value
                if t_auxiliary: auxiliary = t_auxiliary
                snum = sheet.cell(row=i, column=4).value
                num = str(num)[1:]
                col += 1
                row = [auxiliary, snum, 'Модем', num, addr]
                #writer.writerow(row)
                conn = sqlite3.connect(db_filename)
                cursor = conn.cursor()
                sql = f"""INSERT INTO {db_table}(num_tel, purpose, set_addr, aux, type, s_num, date)
                        VALUES (7{row[3]}, '{purpose}', '{addr}', '{auxiliary}', '{row[2]}', '{row[1]}', '{date}')"""
                if log:
                    #ic(sql)
                    print(f"{col} {sql}")
                    #ic(row)
                    #ic(f"{col} {sql}")
                else:
                    cursor.execute(sql)
                    conn.commit()
    cursor.close()
    conn.close()


def mts_on_astra_to_csv_db(excel_filename: str = "", db_filename: str = "", db_table:str = "",
                                     columns_to_readwrite: list = ["", ]):
    date = "2021-06-15"
    purpose = "CSD НА СЧЕТЧИКАХ"
    type = "Модем"
    cwd = os.getcwd()
    print(cwd)
    os.chdir("C:/Users/shamin.a/PycharmProjects/simcard")
    file = 'sim-astra-asque-meter_15_06_2021.csv'
    wb = load_workbook('SIM-карты ГЛОНАСС АСКУЭ LoRa.xlsx')
    print(wb.sheetnames)
    sheet = wb['АСКУЭ. Астра']
    print(sheet.title)

    col = 0
    with open(file, "w", newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file, delimiter=';')
        writer.writerow(['Наименование устройства',             # обязательно
                         'Идентификационный номер',             # необязательно
                         'Тип устройства',                      # обязательно
                         'Номер телефона',                      # обязательно
                         'Адрес, где находится устройство'])    # необязательно
        for i in range(2, 8500):
            num = sheet.cell(row=i, column=4).value
            if num:
                taddr = sheet.cell(row=i, column=2).value
                if taddr: addr = taddr
                snum = sheet.cell(row=i, column=3).value
                t_auxiliary = sheet.cell(row=i, column=5).value
                if t_auxiliary: auxiliary = f"{t_auxiliary} {snum}"
                num = str(num)[1:]
                col += 1
                row = [auxiliary, snum, 'Модем', num, addr]
                writer.writerow(row)
                conn = sqlite3.connect(db_filename)
                cursor = conn.cursor()
                sql = f"""INSERT INTO {db_table}(num_tel, purpose, set_addr, aux, type, s_num, date)
                VALUES (7{row[3]}, '{purpose}', '{addr}', '{auxiliary}', '{row[2]}', '{row[1]}', '{date}')"""
                #print(sql)
                cursor.execute(sql)
                #print(f"{col} {row}")
                conn.commit()
    cursor.close()
    conn.close()
    pass

def mts_on_server_modem_and_region_to_csv_db(excel_filename: str = "", db_filename: str = "", db_table:str = "",
                                     columns_to_readwrite: list = ["", ]):
    date = "2021-06-15"
    purpose = "ОПРОС CSD СЧЕТЧИКОВ"
    type = "Модем"
    cwd = os.getcwd()
    print(cwd)
    os.chdir("C:/Users/shamin.a/PycharmProjects/simcard")
    file = 'sim-server-arm-asque_15_06_2021.csv'
    wb = load_workbook('SIM-карты ГЛОНАСС АСКУЭ LoRa.xlsx')
    print(wb.sheetnames)
    sheet = wb['АСКУЭ. Сервер и АРМ']
    print(sheet.title)
    col = 0
    with open(file, "w", newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file, delimiter=';')
        writer.writerow(['Наименование устройства',
                         'Идентификационный номер',
                         'Тип устройства',
                         'Номер телефона',
                         'Адрес, где находится устройство'])
        for i in range(2, 8500):
            num = sheet.cell(row=i, column=4).value
            if num:
                taddr = sheet.cell(row=i, column=6).value
                if taddr: addr = taddr
                t_auxiliary = sheet.cell(row=i, column=3).value
                if t_auxiliary: auxiliary = t_auxiliary

                snum = ""
                num = str(num)[1:]
                col += 1
                address = f'{addr}, Станке Димитрова, 5В' if addr == "Брянск" else addr
                row = [f"{auxiliary}, модем {col}", snum, 'Модем', num, address]
                writer.writerow(row)
                conn = sqlite3.connect(db_filename)
                cursor = conn.cursor()
                sql = f"""INSERT INTO {db_table}(num_tel, purpose, set_addr, aux, type, s_num, date)
                VALUES (7{row[3]}, '{purpose}', '{address}', '{f"{auxiliary}, модем {col}"}', '{row[2]}', '{row[1]}', '{date}')"""
                #print(sql)
                cursor.execute(sql)
                #print(f"{col} {row}")
                conn.commit()
    cursor.close()
    conn.close()

def mts_on_glonass_vehicle_to_csv_db(excel_filename: str = "", db_filename: str = "", db_table:str = "",
                                     columns_to_readwrite: list = ["", ]):
    date = "2021-06-15"
    purpose = "ГЛОНАСС/GPS АВТОМОБИЛИ"
    type = "Модем"
    cwd = os.getcwd()
    #print(cwd)
    os.chdir("C:/Users/shamin.a/PycharmProjects/simcard")
    file = 'sim-glonass-vehicle_15_06_2021.csv'
    wb = load_workbook('SIM-карты ГЛОНАСС АСКУЭ LoRa.xlsx')
    #print(wb.sheetnames)
    sheet = wb['ГЛОНАСС. Автомобили']
    #print(sheet.title)

    col = 0
    with open(file, "w", newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file, delimiter=';')
        writer.writerow(['Наименование устройства',
                         'Идентификационный номер',
                         'Тип устройства',
                         'Номер телефона',
                         'Адрес, где находится устройство'])
        for i in range(2, 8500):
            num = sheet.cell(row=i, column=6).value
            addr2 = ""
            if num:
                t_auxiliary = sheet.cell(row=i, column=4).value
                if t_auxiliary: auxiliary = f"{t_auxiliary} {col}"
                snum = sheet.cell(row=i, column=5).value
                taddr1 = sheet.cell(row=i, column=1).value
                if taddr1: addr1 = taddr1
                taddr2 = sheet.cell(row=i, column=2).value
                if taddr2: addr2 = taddr2
                addr = f"{addr1} {addr2} {sheet.cell(row=i, column=3).value}".replace('"', '')
                #print(addr)
                num = str(num)[1:]
                col += 1
                row = [auxiliary, snum, 'Модем', num, addr]
                writer.writerow(row)
                conn = sqlite3.connect(db_filename)
                cursor = conn.cursor()
                sql = f"""INSERT INTO {db_table}(num_tel, purpose, set_addr, aux, type, s_num, date)
                VALUES (7{row[3]}, '{purpose}', '{addr}', '{auxiliary}', '{row[2]}', '{row[1]}', '{date}')"""
                #print(sql)
                cursor.execute(sql)
                #print(f"{col} {row}")
                conn.commit()
    cursor.close()
    conn.close()

def mts_glonass_spare_to_csv_db(excel_filename: str = "", db_filename: str = "", db_table:str = "",
                                     columns_to_readwrite: list = ["", ]):
    date = "2021-06-15"
    purpose = "ГЛОНАСС/GPS ЗАПАСНЫЕ"
    type = "Прочее"
    cwd = os.getcwd()
    #print(cwd)
    os.chdir("C:/Users/shamin.a/PycharmProjects/simcard")
    file = 'sim-glonass-spare_15_06_2021.csv'
    wb = load_workbook('SIM-карты ГЛОНАСС АСКУЭ LoRa.xlsx')
    #print(wb.sheetnames)
    sheet = wb['ГЛОНАСС. Запасные']
    #print(sheet.title)
    addr = "Брянск Станке Димитрова 5В"
    tauxiliary = "Запасная СИМ карта"
    col = 0
    with open(file, "w", newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file, delimiter=';')
        writer.writerow(['Наименование устройства',
                         'Идентификационный номер',
                         'Тип устройства',
                         'Номер телефона',
                         'Адрес, где находится устройство'])
        for i in range(5, 8500):
            num = sheet.cell(row=i, column=6).value
            if num:
                snum = sheet.cell(row=i, column=4).value
                if snum is None: snum = ''
                tnum_sim = sheet.cell(row=i, column=5).value
                num_sim = int(tnum_sim.replace('-', '')) if tnum_sim is not None else 0
                num = str(num)[1:]
                auxiliary = f"{tauxiliary} {col}"
                col += 1
                row = [auxiliary, snum, type, num, addr]
                writer.writerow(row)
                conn = sqlite3.connect(db_filename)
                cursor = conn.cursor()
                sql = f"""INSERT INTO {db_table}(num_tel, num_sim, purpose, set_addr, aux, type, s_num, date)
                VALUES (7{row[3]}, {num_sim}, '{purpose}', '{addr}', '{auxiliary}', '{row[2]}', '{row[1]}', '{date}')"""
                #print(sql)
                cursor.execute(sql)
                #print(f"{col} {row}")
                conn.commit()
    cursor.close()
    conn.close()

def mts_netping_to_csv_db(excel_filename: str = "", db_filename: str = "", db_table:str = "",
                                     columns_to_readwrite: list = ["", ]):
    date = "2021-06-15"
    purpose = "СЕТЕВОЙ МОНИТОРИНГ"
    type = "Прочее"
    cwd = os.getcwd()
    #print(cwd)
    os.chdir("C:/Users/shamin.a/PycharmProjects/simcard")
    file = 'sim-netping_15_06_2021.csv'
    wb = load_workbook('SIM-карты ГЛОНАСС АСКУЭ LoRa.xlsx')
    #print(wb.sheetnames)
    sheet = wb['NetPing']
    #print(sheet.title)
    addr = "Брянск Станке Димитрова 5В"
    tauxiliary = "Мониторинг серверных помещений"
    col = 0
    with open(file, "w", newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file, delimiter=';')
        writer.writerow(['Наименование устройства',
                         'Идентификационный номер',
                         'Тип устройства',
                         'Номер телефона',
                         'Адрес, где находится устройство'])
        for i in range(5, 8500):
            num = sheet.cell(row=i, column=3).value
            if num:
                snum = sheet.cell(row=i, column=4).value
                if snum is None: snum = ''
                tnum_sim = sheet.cell(row=i, column=5).value
                num_sim = int(tnum_sim.replace('-', '')) if tnum_sim is not None else 0
                num = str(num)[1:]
                auxiliary = f"{tauxiliary} {col}"
                col += 1
                row = [auxiliary, snum, type, num, addr]
                writer.writerow(row)
                conn = sqlite3.connect(db_filename)
                cursor = conn.cursor()
                sql = f"""INSERT INTO {db_table}(num_tel, num_sim, purpose, set_addr, aux, type, s_num, date)
                VALUES (7{row[3]}, {num_sim}, '{purpose}', '{addr}', '{auxiliary}', '{row[2]}', '{row[1]}', '{date}')"""
                #print(sql)
                cursor.execute(sql)
                #print(f"{col} {row}")
                conn.commit()
    cursor.close()
    conn.close()


def megafon_site(excel_filename: str, db_filename: str, columns_to_readwrite: list):
    wb = load_workbook(MEGAFONFIRST)
    #print(wb.sheetnames)
    sheet = wb['Мобильная связь']
    #print(sheet.title)
    col = 1
    date = '15.06.2021'
    opsos = 'MEGAFON'
    for i in range(2, 9500):
        num = sheet.cell(row=i, column=1).value
        if num:
            row = [col, num, date, opsos, ]
            for c in columns_to_readwrite:
                tval = sheet.cell(row=i, column=c).value
                if tval:
                    val = tval
                row.append(tval)
            col += 1
            print(row)

def megafon_to_csv(excel_filename: str = "", db_filename: str = "", columns_to_readwrite: list = ["", ]):
    cwd = os.getcwd()
    print(cwd)
    os.chdir("C:/Users/shamin.a/PycharmProjects/simcard")
    file = 'sim-megafon.csv'
    wb = load_workbook('megafon_sim.xlsx')
    print(wb.sheetnames)
    sheet = wb['telemeh']
    print(sheet.title)
    col = 0
    with open(file, "w", newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file, delimiter=';')
        writer.writerow(['Наименование устройства',
                         'Идентификационный номер',
                         'Тип устройства',
                         'Номер телефона',
                         'Адрес, где находится устройство'])
        for i in range(2, 8500):
            num = sheet.cell(row=i, column=1).value
            if num:
                taddr = sheet.cell(row=i, column=3).value
                if taddr:
                    addr = taddr
                    auxiliary = "Телемеханика " + addr
                #if t_auxiliary: auxiliary = t_auxiliary
                snum = sheet.cell(row=i, column=2).value
                num = str(num)[1:]
                col += 1
                writer.writerow([auxiliary, snum, 'Модем', num, "Брянск, " + addr])


def tele2_site_to_csv_db(excel_filename: str = "", db_filename: str = "", db_table:str = "",
                                     columns_to_readwrite: list = ["", ]):
    date = "2021-06-15"
    purpose = "LORA"
    type = "Модем"
    cwd = os.getcwd()
    #print(cwd)
    os.chdir("C:/Users/shamin.a/PycharmProjects/simcard")
    file = 'sim-tele2_15_06_2021.csv'
    wb = load_workbook('tele2_M2M_lk_sim_15_06_2021.xlsx')
    #print(wb.sheetnames)
    sheet = wb['Лист1']
    #print(sheet.title)

    col = 0
    with open(file, "w", newline='', encoding='utf-8') as csv_file:
        writer = csv.writer(csv_file, delimiter=';')
        writer.writerow(['Наименование устройства',
                         'Идентификационный номер',
                         'Тип устройства',
                         'Номер телефона',
                         'Адрес, где находится устройство'])
        for i in range(2, 8500):
            num = sheet.cell(row=i, column=1).value
            if num:
                taddr = sheet.cell(row=i, column=4).value
                if taddr: addr = taddr
                t_auxiliary = sheet.cell(row=i, column=2).value
                if t_auxiliary: auxiliary = t_auxiliary
                t_type = sheet.cell(row=i, column=5).value
                if t_type: type = t_type
                if auxiliary[:6] == "Резерв":
                    purpose = "LORA/Телемеханика запасные"
                elif auxiliary[-10:] == "мониторинг":
                    purpose = "СЕТЕВОЙ МОНИТОРИНГ"
                elif auxiliary[:4] == "LoRa":
                    purpose = "LORA"
                elif auxiliary[:2] == "ЛК":
                    purpose = "Личный кабинет"
                elif auxiliary[-9:] == "Интернета":
                    purpose = "Интернет"
                else:
                    purpose = "Телемеханика"
                snum = sheet.cell(row=i, column=3).value
                num = str(num)[2:].strip().replace("-", "").replace(" ", "")
                col += 1
                row = [auxiliary, snum, type, num, addr]
                writer.writerow(row)
                conn = sqlite3.connect(db_filename)
                cursor = conn.cursor()
                sql = f"""INSERT INTO {db_table}(num_tel, purpose, set_addr, aux, type, s_num, date)
                VALUES (7{row[3]}, '{purpose}', '{addr}', '{auxiliary}', '{row[2]}', '{row[1]}', '{date}')"""
                print(sql)
                #cursor.execute(sql)
                print(f"{col} {row}")
                conn.commit()
    cursor.close()
    conn.close()



if __name__ == '__main__':
    cwd = os.getcwd()
    #print(cwd)
    os.chdir(".")
    make_db()
    if READFIRSTMTS:
        mts_site_to_db("charge_report_electric_energy_meters_15_06_2021.xlsx", "Charges", "simdatabase.db", [6, ])
        mts_site_to_db("charge_report_IoT_gprs_sms_call_23_07_2021.xlsx", "Charges", "simdatabase.db", [6, ])
        mts_site_to_db("charge_report_gps_23_07_2021.xlsx", "Charges", "simdatabase.db", [6, ])
        mts_site_to_db("charge_report_management_23_07_2021.xlsx", "Charges", "simdatabase.db", [6, ])
        mts_site_to_db("charge_report_request_for_electricity_meters_23_07_2021.xlsx", "Charges", "simdatabase.db",
                       [6, ])
        mts_on_piramida_server_to_csv_db("sim-piramida-asque-meter_15_06_2021.csv", "simdatabase.db", "mts_current")
        mts_on_astra_to_csv_db("sim-astra-asque-meter_15_06_2021.csv", "simdatabase.db", "mts_current")
        mts_on_server_modem_and_region_to_csv_db("sim-server-arm-asque_15_06_2021.csv", "simdatabase.db", "mts_current")
        mts_on_glonass_vehicle_to_csv_db("sim-glonass-vehicle_15_06_2021.csv", "simdatabase.db", "mts_current")
        mts_glonass_spare_to_csv_db("sim-glonass-spare_15_06_2021.csv", "simdatabase.db", "mts_current")
        mts_netping_to_csv_db("sim-netping_15_06_2021.csv", "simdatabase.db", "mts_current")
    if READFIRSTMTS210813:
        mts_site_to_db("charge_report_13_08_2021.xlsx", "Charges", "simdatabase.db", [6, ], '2021-08-13')
    if READFIRSTMEGAFON:
        MEGAFONFIRST: str = 'mobileSubscribers_20210723_140421_15_06_2021.xlsx'
        megafon_site(MEGAFONFIRST, 'simdatabase.db', [7, ])
    if READFIRSTMTS210910:
        mts_site_to_db("charge_report_electric_energy_meters_10_09_2021.xlsx", "Charges",
                       "simdatabase.db", [6, ], '2021-09-10')
    #megafon_to_csv()
    #tele2_site_to_csv_db("sim-tele2_15_06_2021.csv", "simdatabase.db", "tele2_current")
    #mts_site210813_minus_current210615_to_csv()
    #mts_site_charge_report_electric_energy_meters_15_06_2021_to_csv()
    #mts_site_now_minus_previous_to_csv()
    if READFIRSTMTS210910:
        mts_site_to_db("charge_report_electric_energy_meters_10_09_2021.xlsx", "Charges",
                       "simdatabase.db", [6, ], '2021-09-10')
    if READFIRSTMTS211013:
        mts_site_to_db("charge_report_electric_energy_meters_13_10_2021.xlsx", "Charges",
                       "simdatabase.db", [6, ], '2021-10-13', log = True)
    if READFIRSTMTS211027:
        mts_site_to_db("charge_report_electric_energy_meters_27_10_2021.xlsx", "Charges",
                       "simdatabase.db", [6, ], '2021-10-27', log = True)

    #mts_site_now_minus_previous_to_csv(now='2021-10-27', previous_date='2021-10-13',
    #                                   db_filename = "simdatabase.db", start_count =10401)

    mts_on_piramida_server_to_db("2021-11-25", "SIM-карты АСКУЭ Пирамида 211125.xlsm", "simdatabase.db", "mts_current",
                                 log=True)
